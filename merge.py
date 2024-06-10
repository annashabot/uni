import csv
from openpyxl import Workbook
from typing import NoReturn
from openpyxl.styles import numbers

arr=["read_stat","alpha-diversity","phyla", "class", "order", "family", "genus", "species"]
wb = Workbook()
ws0 = wb.active
numrow=[]
with open('read_stat.xlsx') as f:
	reader = csv.reader(f, delimiter="\t")
	for row in reader:
    	for cell in row:
        	try:
            	numrow.append(float(cell))
        	except:
            	numrow.append(cell)
    	ws0.append(numrow)
    	print(numrow)
    	numrow=[]
ws0.title = "read_stat"
ws1 = wb.create_sheet()
with open('alpha_diversity.xlsx') as f:
	reader = csv.reader(f, delimiter='\t')
	for row in reader:
    	for cell in row:
        	try:
            	numrow.append(float(cell))
        	except:
            	numrow.append(cell)
    	ws1.append(numrow)
    	print(numrow)
    	numrow=[]
ws1.title = "alpha-diversity"
ws2 = wb.create_sheet()
with open('phyla.xlsx') as f:
	reader = csv.reader(f, delimiter='\t')
	for row in reader:
    	for cell in row:
        	try:
            	numrow.append(float(cell))
        	except:
            	numrow.append(cell)
    	ws2.append(numrow)
    	print(numrow)
    	numrow=[]
ws2.title = "phyla"
ws3 = wb.create_sheet()
with open('class.xlsx') as f:
	reader = csv.reader(f, delimiter='\t')
	for row in reader:
    	for cell in row:
        	try:
            	numrow.append(float(cell))
        	except:
            	numrow.append(cell)
    	ws3.append(numrow)
    	print(numrow)
    	numrow=[]
ws3.title = "class"
ws4 = wb.create_sheet()
with open('order.xlsx') as f:
	reader = csv.reader(f, delimiter='\t')
	for row in reader:
    	for cell in row:
        	try:
            	numrow.append(float(cell))
        	except:
            	numrow.append(cell)
    	ws4.append(numrow)
    	print(numrow)
    	numrow=[]
ws4.title = "order"
ws5 = wb.create_sheet()
with open('family.xlsx') as f:
	reader = csv.reader(f, delimiter='\t')
	for row in reader:
    	for cell in row:
        	try:
            	numrow.append(float(cell))
        	except:
            	numrow.append(cell)
    	ws5.append(numrow)
    	print(numrow)
    	numrow=[]
ws5.title = "family"
ws6 = wb.create_sheet()
with open('genus.xlsx') as f:
	reader = csv.reader(f, delimiter='\t')
	for row in reader:
    	for cell in row:
        	try:
            	numrow.append(float(cell))
        	except:
            	numrow.append(cell)
    	ws6.append(numrow)
    	print(numrow)
    	numrow=[]
ws6.title = "genus"
ws7 = wb.create_sheet()
with open('species.xlsx') as f:
	reader = csv.reader(f, delimiter='\t')
	for row in reader:
    	for cell in row:
        	try:
            	numrow.append(float(cell))
        	except:
            	numrow.append(cell)
    	ws7.append(numrow)
    	print(numrow)
    	numrow=[]
ws7.title = "species"
wb.save('report.xlsx')
